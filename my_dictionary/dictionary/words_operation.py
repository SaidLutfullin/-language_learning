import os
from datetime import date, timedelta
from random import randint

from django.conf import settings
from django.db.models import Q
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import Words


def get_question(user, excluded_word_id):
    try:
        today_date = date.today()
        words = Words.objects.filter(user_id=user.id, language=user.language_learned)
        if excluded_word_id is not None:
            words = words.exclude(pk=excluded_word_id)
        word = words.filter(asking_date=today_date).order_by("?").first()
        if word is None:
            word = words.filter(asking_date__lt=today_date).order_by("?").first()
    except Exception as error:
        word = None
        logger.error(f"Functions get_word. Error {str(error)}")
    finally:
        return word


def send_answer(user, word_id, is_correct):
    word = Words.objects.filter(pk=word_id, user_id=user.id).first()
    # changing box number
    if is_correct:
        if word.box_number != 6:
            word.box_number = word.box_number + 1
    else:
        word.box_number = 0
    # generating asking date
    today_date = date.today()
    if word.box_number == 0:
        interval = timedelta(days=1)
    elif word.box_number == 1:
        interval = timedelta(days=2)
    elif word.box_number == 2:
        interval = timedelta(days=3)
    elif word.box_number == 3:
        interval = timedelta(days=7)
    elif word.box_number == 4:
        interval = timedelta(days=14)
    elif word.box_number == 5:
        interval = timedelta(days=30)
    elif word.box_number == 6:
        interval = timedelta(days=randint(30, 40))
    word.asking_date = today_date + interval
    word.save()


@logger.catch
def get_words_list(user, search_query):
    words_list = Words.objects.filter(user_id=user.id, language=user.language_learned)
    search_query = search_query.replace(" ", "")
    if search_query != "":
        words_list = words_list.filter(
            Q(russian_word__icontains=search_query)
            | Q(foreign_word__icontains=search_query)
            | Q(context__icontains=search_query)
        )
    return words_list


@logger.catch
def get_dictionary_statistics(user):
    words_list = Words.objects.filter(user_id=user.pk, language=user.language_learned)
    statistics = {
        "words_total_number": words_list.count(),
        "words_for_today_number": words_list.filter(
            asking_date__lte=date.today()
        ).count(),
        "learned_words_number": words_list.filter(box_number=6).count(),
    }
    return statistics


def export_words(user, fields):
    words_list = Words.objects.filter(
        user_id=user.id, language=user.language_learned
    ).values(*fields)

    work_book = Workbook()
    spreadsheet = work_book.active
    spreadsheet.title = "Мой словарь"
    bd = Side(border_style="thin")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    fields_in_russian = {
        "russian_word": "Русское слово",
        "foreign_word": "Иностранное слово",
        "context": "Контекст",
        "box_number": "Коробка",
        "asking_date": "Дата повторения",
    }
    fields = [fields_in_russian[field_name] for field_name in fields]
    spreadsheet.append(fields)

    for cell in spreadsheet[1]:
        column_letter = get_column_letter(cell.column)
        spreadsheet.column_dimensions[column_letter].width = 25
        cell.border = border
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    spreadsheet.row_dimensions[1].height = 25

    for word in words_list:
        row_number = spreadsheet.max_row + 1
        for column_index, column_key in enumerate(word):
            column_letter = get_column_letter(column_index + 1)
            if column_key == "asking_date":
                spreadsheet[f"{column_letter}{row_number}"] = word[column_key].strftime(
                    "%d.%m.%Y"
                )
            else:
                spreadsheet[f"{column_letter}{row_number}"] = word[column_key]

            if column_key == "asking_date" or column_key == "box_number":
                spreadsheet[f"{column_letter}{row_number}"].alignment = Alignment(
                    wrap_text=True, horizontal="center"
                )
            else:
                spreadsheet[f"{column_letter}{row_number}"].alignment = Alignment(
                    wrap_text=True
                )
            spreadsheet[f"{column_letter}{row_number}"].border = border
            spreadsheet[f"{column_letter}{row_number}"].font = Font(size=10)
    path = os.path.join(settings.MEDIA_ROOT, "users", str(user.id), "export.xlsx")
    work_book.save(path)
    return path
